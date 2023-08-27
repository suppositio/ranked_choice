import tkinter as tk
from tkinter import messagebox as mb
from openpyxl import load_workbook
from os.path import splitext, basename

from .path_editor import PathEditor
from .name_editor import NameEditor
from .slate_column import SlateColumn
from .ranking_column import RankingColumn

class VotingLoader:
    def __init__(self, master):
        self.__master = master

    @property
    def result(self):
        self.__load_path = self.__get_load_path()
        if not self.__load_path:
            return None
        
        while not self.__read_data():
            if self.__confirm_retry():
                self.__load_path = self.__get_load_path()
                if not self.__load_path:
                    return None
            else:
                return None

        if not self.__set_voting_name_if_needed():
            return None

        self.__create_voting()
        return self.__voting

    def __get_load_path(self):
        load_path_editor = PathEditor(self.__master, "Excel", "xlsx", PathEditor.LOAD)
        return load_path_editor.result
    
    def __read_data(self):
        try:
            workbook = load_workbook(self.__load_path)
            sheet = workbook.active
            self.__data_array = []

            for row in sheet.iter_rows(values_only = True):
                self.__data_array.append(list(row))

            if not self.__validate_data(self.__data_array):
                return False
            
            return True
        
        except Exception:
            return False

    @staticmethod    
    def __validate_data(data):
        if not data or len(data) < 2:
            return False
        
        column_label_row = data[0]
        if not all(column_label and isinstance(column_label, str) for column_label in column_label_row[1:]):
            return False

        benchmark_length = len(column_label_row)
        previous_value_row_labels = []
        for value_row in data[1:]:
            if len(value_row) != benchmark_length:
                return False
            value_row_label = value_row[0]
            if not (value_row_label and isinstance(value_row_label, str)):
                return False
            if value_row_label in previous_value_row_labels:
                return False
            previous_value_row_labels.append(value_row_label)
            if len(value_row) > 1:
                if not all(value and isinstance(value, int) for value in value_row[1:]):
                    return False

        return True

    def __set_voting_name_if_needed(self):
        if not self.__data_array[0][0]:
            default_voting_name = splitext(basename(self.__load_path))[0]
            voting_name_editor = NameEditor(self.__master, "Введіть назву голосування!", default_voting_name)
            if not (voting_name := voting_name_editor.result):
                return False
            self.__data_array[0][0] = voting_name

        return True

    def __create_voting(self):
        self.__voting = tk.Frame(self.__master)

        voting_name = self.__data_array[0][0]
        voting_slate = [row[0] for row in self.__data_array[1:]]
        slate_column = SlateColumn(self.__voting, voting_name, voting_slate)
        slate_column.pack(side = tk.LEFT, anchor = tk.N)

        if len(self.__data_array[0]) > 1:
            for value_column in list(map(list, zip(*self.__data_array)))[1:]:
                voter = value_column[0]
                vote = {voter:rank for voter, rank in zip(voting_slate, value_column[1:])}
                ranking_column = RankingColumn(self.__voting, voting_slate, voter, vote)
                ranking_column.pack(side = tk.LEFT, anchor = tk.N)

    def __confirm_retry(self):
        return mb.askokcancel(message = "Не вдалося завантажити файл.\nСпробувати інший?")
